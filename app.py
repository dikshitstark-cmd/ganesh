"""
app.py - Ganesh Utsav Receipt Management System
Flask application entry point. Run with:  python app.py

Environment variables (all optional, sensible defaults for local dev):
  SECRET_KEY       Flask session signing key. SET THIS for any live deployment.
  ADMIN_USERNAME   Username for the auto-created first admin account (default: admin)
  ADMIN_PASSWORD   Password for that account (default: admin123 -- CHANGE IT)
  DATABASE_PATH    Path to the SQLite file (default: receipts.db next to this file)
  PORT             Port to listen on when run directly with `python app.py`
"""

import csv
import io
import os
import zipfile
from datetime import datetime
from functools import wraps
from pathlib import Path

import openpyxl
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, abort, session
)
from werkzeug.security import check_password_hash

import db
import utils

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
GENERATED_DIR = BASE_DIR / "generated"
UPLOAD_DIR.mkdir(exist_ok=True)
GENERATED_DIR.mkdir(exist_ok=True)

ALLOWED_MODES = ["Cash", "UPI", "Cash/UPI", "Bank Transfer", "Cheque", "Other"]
ALLOWED_EXCEL_EXT = {".xlsx", ".xls"}

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret-key-in-production")
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB upload cap


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        if session.get("role") != "admin":
            abort(403)
        return view(*args, **kwargs)
    return wrapped


@app.context_processor
def inject_current_user():
    return {
        "current_username": session.get("username"),
        "current_role": session.get("role"),
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))

    if request.method == "GET":
        return render_template("login.html", error=None)

    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    user = db.get_user_by_username(username)

    if not user or not check_password_hash(user["password_hash"], password):
        return render_template("login.html", error="Invalid username or password."), 401

    session.clear()
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    session["role"] = user["role"]
    next_url = request.form.get("next") or url_for("dashboard")
    return redirect(next_url)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/account/password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "GET":
        return render_template("change_password.html", error=None)

    current = request.form.get("current_password", "")
    new = request.form.get("new_password", "")
    confirm = request.form.get("confirm_password", "")
    user = db.get_user(session["user_id"])

    if not check_password_hash(user["password_hash"], current):
        return render_template("change_password.html", error="Current password is incorrect."), 400
    if len(new) < 6:
        return render_template("change_password.html", error="New password must be at least 6 characters."), 400
    if new != confirm:
        return render_template("change_password.html", error="New password and confirmation do not match."), 400

    db.update_user_password(user["id"], new)
    flash("Password updated successfully.", "success")
    return redirect(url_for("dashboard"))


@app.route("/admin/users", methods=["GET", "POST"])
@admin_required
def admin_users():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        if not username or len(username) < 3:
            flash("Username must be at least 3 characters.", "error")
        elif len(password) < 6:
            flash("Password must be at least 6 characters.", "error")
        elif db.get_user_by_username(username):
            flash("That username is already taken.", "error")
        else:
            db.create_user(username, password, role="subadmin")
            flash(f"Sub-admin '{username}' created.", "success")
        return redirect(url_for("admin_users"))

    users = db.list_users()
    return render_template("admin_users.html", users=users)


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def delete_subadmin(user_id):
    target = db.get_user(user_id)
    if not target:
        abort(404)
    if target["role"] == "admin":
        flash("Admin accounts cannot be deleted from here.", "error")
    elif target["id"] == session.get("user_id"):
        flash("You cannot delete your own account while logged in.", "error")
    else:
        db.delete_user(user_id)
        flash(f"Sub-admin '{target['username']}' removed.", "success")
    return redirect(url_for("admin_users"))


@app.errorhandler(403)
def forbidden(e):
    return render_template("error.html", code=403, message="You don't have permission to view this page."), 403


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def build_receipt_context(row):
    """Turn a sqlite Row into everything the receipt template needs."""
    r = dict(row)
    upi_id = db.get_setting("upi_id")
    org_en = db.get_setting("org_name_en")
    org_te = db.get_setting("org_name_te")

    qr_data_uri = None
    if r["due_amount"] > 0 and r.get("show_qr", 1):
        qr_data_uri = utils.generate_upi_qr_data_uri(
            upi_id=upi_id,
            payee_name=org_en,
            amount=r["due_amount"],
            note=f"{r['purpose']} {r['receipt_no']}",
            txn_ref=r["receipt_no"],
        )

    r["amount_words"] = utils.amount_in_words(r["total_amount"])
    r["qr_data_uri"] = qr_data_uri
    r["org_name_en"] = org_en
    r["org_name_te"] = org_te
    return r


def render_receipt_html(row, for_pdf=False):
    """for_pdf=True renders without the remote Google Fonts <link>. Some
    WeasyPrint/fontTools combinations crash when subsetting certain
    variable webfonts (see README troubleshooting section) -- rendering
    against the locally-installed static Noto Sans Telugu / system fonts
    avoids that entirely and looks effectively identical."""
    return render_template(
        "receipt.html", r=build_receipt_context(row),
        standalone=True, for_pdf=for_pdf,
    )


def parse_id_list(raw):
    """Parse a comma-separated or repeated-field list of ints, silently
    dropping anything non-numeric."""
    out = []
    for part in raw:
        for piece in str(part).split(","):
            piece = piece.strip()
            if piece.isdigit():
                out.append(int(piece))
    return out


# ---------------------------------------------------------------------------
# Core pages
# ---------------------------------------------------------------------------

@app.route("/")
@login_required
def home():
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
@login_required
def dashboard():
    search = request.args.get("q", "").strip()
    status = request.args.get("status", "All")
    receipts = db.list_receipts(search=search or None, status=status)

    totals = {
        "count": len(receipts),
        "total": sum(r["total_amount"] for r in receipts),
        "paid": sum(r["paid_amount"] for r in receipts),
        "due": sum(r["due_amount"] for r in receipts),
    }
    return render_template(
        "dashboard.html",
        receipts=receipts,
        search=search,
        status=status,
        totals=totals,
        statuses=["All", "Paid", "Partial", "Unpaid"],
    )


@app.route("/receipt/new", methods=["GET", "POST"])
@login_required
def new_receipt():
    if request.method == "GET":
        return render_template("new_receipt.html", modes=ALLOWED_MODES, error=None)

    try:
        name = request.form.get("name", "").strip()
        address = request.form.get("address", "").strip()
        phone = request.form.get("phone", "").strip()
        purpose = request.form.get("purpose", "Ganesh Chanda").strip() or "Ganesh Chanda"
        mode = request.form.get("mode", "Cash")
        total_amount = request.form.get("total_amount", "").strip()
        paid_amount = request.form.get("paid_amount", "0").strip() or "0"
        payment_status_choice = request.form.get("payment_status", "unpaid")
        show_qr = request.form.get("show_qr") == "on"

        if not name or not address:
            raise ValueError("Name and Address are required.")
        if mode not in ALLOWED_MODES:
            raise ValueError("Invalid payment mode selected.")

        try:
            total_amount = float(total_amount)
            paid_amount = float(paid_amount)
        except ValueError:
            raise ValueError("Amount fields must be numeric.")

        if total_amount <= 0:
            raise ValueError("Total amount must be greater than zero.")
        if paid_amount < 0 or paid_amount > total_amount:
            raise ValueError("Paid amount cannot be negative or exceed the total amount.")

        # Respect an explicit "mark as paid in full" choice from the form
        if payment_status_choice == "paid":
            paid_amount = total_amount
        elif payment_status_choice == "unpaid":
            paid_amount = 0.0
        # "partial" keeps whatever paid_amount was typed in

        new_id = db.create_receipt(
            name=name, address=address, total_amount=total_amount,
            mode=mode, purpose=purpose, phone=phone or None,
            paid_amount=paid_amount, show_qr=show_qr,
        )
        flash("Receipt created successfully.", "success")
        return redirect(url_for("view_receipt", receipt_id=new_id))

    except ValueError as e:
        return render_template("new_receipt.html", modes=ALLOWED_MODES, error=str(e)), 400
    except Exception:
        return render_template(
            "new_receipt.html", modes=ALLOWED_MODES,
            error="Something went wrong creating the receipt. Please check your input."
        ), 500


@app.route("/receipt/<int:receipt_id>")
@login_required
def view_receipt(receipt_id):
    row = db.get_receipt(receipt_id)
    if not row:
        abort(404)
    context = build_receipt_context(row)
    return render_template("receipt.html", r=context, standalone=False)


@app.route("/receipt/<int:receipt_id>/payment", methods=["POST"])
@login_required
def update_payment(receipt_id):
    row = db.get_receipt(receipt_id)
    if not row:
        abort(404)

    action = request.form.get("action")
    try:
        if action == "mark_paid":
            db.mark_paid(receipt_id)
            flash(f"Receipt {row['receipt_no']} marked as Paid. "
                  f"Reprint it any time to hand over a Paid copy.", "success")
        elif action == "update_partial":
            amount = float(request.form.get("paid_amount", "0"))
            if amount < 0 or amount > row["total_amount"]:
                raise ValueError("Paid amount must be between 0 and the total amount.")
            db.update_payment(receipt_id, amount)
            flash(f"Payment updated for receipt {row['receipt_no']}.", "success")
        else:
            raise ValueError("Unknown action.")
    except ValueError as e:
        flash(str(e), "error")

    return redirect(request.referrer or url_for("dashboard"))


@app.route("/receipt/<int:receipt_id>/toggle-qr", methods=["POST"])
@login_required
def toggle_qr(receipt_id):
    row = db.get_receipt(receipt_id)
    if not row:
        abort(404)
    new_state = not bool(row["show_qr"])
    db.set_show_qr(receipt_id, new_state)
    flash(f"QR code {'enabled' if new_state else 'removed'} for receipt {row['receipt_no']}.", "success")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/receipt/<int:receipt_id>/delete", methods=["POST"])
@login_required
def delete_receipt(receipt_id):
    row = db.get_receipt(receipt_id)
    if not row:
        abort(404)
    db.delete_receipt(receipt_id)
    flash(f"Receipt {row['receipt_no']} deleted.", "success")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/receipts/delete-selected", methods=["POST"])
@login_required
def delete_selected():
    ids = parse_id_list(request.form.getlist("ids"))
    if not ids:
        flash("No receipts were selected.", "error")
        return redirect(request.referrer or url_for("dashboard"))
    db.delete_receipts(ids)
    flash(f"Deleted {len(ids)} receipt(s).", "success")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/receipts/print-selected")
@login_required
def print_selected():
    ids = parse_id_list(request.args.getlist("ids"))
    if not ids:
        flash("Select at least one receipt to print.", "error")
        return redirect(url_for("dashboard"))
    rows = db.get_receipts_by_ids(ids)
    receipts = [build_receipt_context(r) for r in rows]
    return render_template("print_selected.html", receipts=receipts)


@app.route("/receipt/<int:receipt_id>/whatsapp")
@login_required
def whatsapp_receipt(receipt_id):
    row = db.get_receipt(receipt_id)
    if not row:
        abort(404)
    receipt_url = url_for("view_receipt", receipt_id=receipt_id, _external=True)
    message = (
        f"Namaste {row['name']},\n"
        f"Thank you for your contribution of Rs. {row['total_amount']:.2f} "
        f"({row['status']}) towards {row['purpose']}.\n"
        f"Receipt No: {row['receipt_no']}\n"
        f"View/download your receipt here: {receipt_url}"
    )
    link = utils.whatsapp_link(row["phone"], message)
    if not link:
        flash("No phone number saved for this receipt, so a WhatsApp link "
              "could not be generated. Add a phone number to the receipt "
              "record first (it is never printed on the receipt).", "error")
        return redirect(request.referrer or url_for("dashboard"))
    return redirect(link)


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

EXPORT_COLUMNS = [
    "receipt_no", "date", "name", "address", "phone", "purpose",
    "total_amount", "paid_amount", "due_amount", "mode", "status",
]


@app.route("/export/csv")
@login_required
def export_csv():
    receipts = db.all_receipts_as_dicts()
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for r in receipts:
        writer.writerow(r)
    mem = io.BytesIO(buf.getvalue().encode("utf-8-sig"))  # BOM so Excel reads UTF-8 correctly
    filename = f"ganesh_utsav_receipts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(mem, as_attachment=True, download_name=filename, mimetype="text/csv")


@app.route("/export/excel")
@login_required
def export_excel():
    receipts = db.all_receipts_as_dicts()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipts"

    headers = ["Receipt No", "Date", "Name", "Address", "Phone", "Purpose",
               "Total Amount", "Paid Amount", "Due Amount", "Mode", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for r in receipts:
        ws.append([r.get(c) for c in EXPORT_COLUMNS])

    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    filename = f"ganesh_utsav_receipts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        mem, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------------------------------------------------------------------
# Bulk generation
# ---------------------------------------------------------------------------

REQUIRED_BULK_COLUMNS = {"name", "address", "amount", "mode"}


@app.route("/bulk", methods=["GET", "POST"])
@login_required
def bulk_upload():
    if request.method == "GET":
        return render_template("bulk_upload.html", error=None, results=None)

    file = request.files.get("excel_file")
    if not file or file.filename == "":
        return render_template("bulk_upload.html", error="Please choose an Excel file.", results=None), 400

    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXCEL_EXT:
        return render_template(
            "bulk_upload.html",
            error="Invalid file type. Please upload a .xlsx or .xls file.",
            results=None), 400

    save_path = UPLOAD_DIR / f"bulk_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
    file.save(save_path)

    try:
        wb = openpyxl.load_workbook(save_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        return render_template(
            "bulk_upload.html",
            error="Could not read the Excel file. Make sure it is a valid, unprotected .xlsx file.",
            results=None), 400

    if not rows:
        return render_template("bulk_upload.html", error="The uploaded file is empty.", results=None), 400

    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    missing = REQUIRED_BULK_COLUMNS - set(header)
    if missing:
        return render_template(
            "bulk_upload.html",
            error=f"Missing required column(s): {', '.join(sorted(missing))}. "
                  f"Expected columns: Name, Address, Amount, Mode.",
            results=None), 400

    col_idx = {h: i for i, h in enumerate(header)}
    created_rows = []
    errors = []

    for line_no, row in enumerate(rows[1:], start=2):
        if row is None or all(c in (None, "") for c in row):
            continue
        try:
            name = str(row[col_idx["name"]]).strip() if row[col_idx["name"]] else ""
            address = str(row[col_idx["address"]]).strip() if row[col_idx["address"]] else ""
            amount_raw = row[col_idx["amount"]]
            mode = str(row[col_idx["mode"]]).strip() if row[col_idx["mode"]] else "Cash"

            if not name or not address:
                raise ValueError("Name/Address missing")
            if amount_raw in (None, ""):
                raise ValueError("Amount missing")
            amount = float(amount_raw)
            if amount <= 0:
                raise ValueError("Amount must be > 0")
            if mode not in ALLOWED_MODES:
                mode = "Other"

            new_id = db.create_receipt(
                name=name, address=address, total_amount=amount,
                mode=mode, paid_amount=0.0,  # bulk-generated receipts start Unpaid
            )
            created_rows.append(new_id)
        except Exception as e:
            errors.append(f"Row {line_no}: {e}")

    if not created_rows:
        return render_template(
            "bulk_upload.html",
            error="No valid rows were found. " + ("; ".join(errors) if errors else ""),
            results=None), 400

    # Generate a ZIP of individual printable HTML receipts (see note in README
    # about optionally rendering these to PDF with WeasyPrint).
    zip_name = f"bulk_receipts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    zip_path = GENERATED_DIR / zip_name
    pdf_engine_ok = True
    try:
        from weasyprint import HTML  # imported lazily; optional dependency
    except Exception:
        pdf_engine_ok = False

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for rid in created_rows:
            row = db.get_receipt(rid)
            html_str = render_receipt_html(row, for_pdf=True)
            safe_name = row["receipt_no"].replace("/", "-")
            if pdf_engine_ok:
                try:
                    pdf_bytes = HTML(string=html_str, base_url=request.url_root).write_pdf()
                    zf.writestr(f"{safe_name}.pdf", pdf_bytes)
                    continue
                except Exception:
                    pass  # fall back to HTML for this receipt
            zf.writestr(f"{safe_name}.html", html_str)

    results = {
        "created": len(created_rows),
        "errors": errors,
        "zip_name": zip_name,
        "pdf_engine_ok": pdf_engine_ok,
        "ids": created_rows,
    }
    return render_template("bulk_upload.html", error=None, results=results)


@app.route("/generated/<path:filename>")
@login_required
def download_generated(filename):
    path = GENERATED_DIR / filename
    if not path.exists():
        abort(404)
    return send_file(path, as_attachment=True)


@app.route("/bulk/sample")
@login_required
def bulk_sample():
    """Serve a ready-made sample Excel file for the bulk upload format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipts"
    ws.append(["Name", "Address", "Amount", "Mode"])
    ws.append(["Ravi Kumar", "12-3, Nallakunta, Hyderabad", 501, "Cash"])
    ws.append(["Lakshmi Devi", "45-6, Ameerpet, Hyderabad", 1001, "UPI"])
    ws.append(["Sai Teja", "78, Kukatpally, Hyderabad", 251, "Cash"])
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem, as_attachment=True, download_name="sample_bulk_receipts.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ---------------------------------------------------------------------------
# Error handlers
# ---------------------------------------------------------------------------

@app.errorhandler(404)
def not_found(e):
    return render_template("error.html", code=404, message="Page or receipt not found."), 404


@app.errorhandler(413)
def too_large(e):
    return render_template("error.html", code=413, message="Uploaded file is too large (max 10 MB)."), 413


@app.errorhandler(500)
def server_error(e):
    return render_template("error.html", code=500, message="An unexpected error occurred."), 500


# ---------------------------------------------------------------------------
# Entry point (local dev). In production a WSGI server (gunicorn) imports
# `app` directly and calls db.init_db() is triggered below at import time
# either way, so both `python app.py` and `gunicorn app:app` work.
# ---------------------------------------------------------------------------

db.init_db()

if __name__ == "__main__":
    # use_reloader=False: the watchdog-based reloader watches the whole
    # working directory by default, which includes uploads/ and generated/.
    # Writing a bulk ZIP or an uploaded Excel file would otherwise trigger
    # a mid-request server restart. Set FLASK_DEBUG=1 env var for verbose
    # tracebacks during development without enabling the reloader.
    debug_mode = os.environ.get("FLASK_DEBUG") == "1"
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=debug_mode, use_reloader=False, host="0.0.0.0", port=port)
